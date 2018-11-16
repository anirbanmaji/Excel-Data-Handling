import openpyxl
wb = openpyxl.load_workbook('fruits.xlsx')
print(wb.sheetnames)
mainSheet = wb['Main']
print(mainSheet['A1'].value)
print()
maxrow = mainSheet.max_row
minrow = mainSheet.min_row
print(maxrow,', ', minrow)
for i in range(1,11):
    print(mainSheet.cell(row=i,column=2).value)

print()
print(openpyxl.cell.cell.column_index_from_string('CD'))
print(openpyxl.cell.cell.get_column_letter(27))


    
